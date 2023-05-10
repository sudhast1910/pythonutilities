import openpyxl
from openpyxl import load_workbook
path = "C:\\Example.xlsx"
#below is sample excel sheet data  - read columnB and if repo matches with dataToMatchinExcelAndUpdate first element then update columnC and columnD with second,third element
#colomnB                           colomnC    colomnD    colomnE
#https://git.example.com/repo15
#https://git.example.com/repo1
#https://git.example.com/repo10

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
print(sheet_obj.title)
first_column = sheet_obj['B']
dataToMatchinExcelAndUpdate=['https://git.example.com/repo1--origin/branch1--2022-03-20 18:07:08','https://git.example.com/repo2--origin/branch1::origin/branch2--2022-03-20 18:07:08::2023-02-10 18:07:08']
for repoName in dataSettoMatchinExcel:
	repoList=repoName.split("--")
# Print the contents
	for row in sheet_obj.iter_rows():
		for cell in row:
			if cell.value == repoList[0]:
				print (repoList[0])				
				print(cell.row)	
        #column 3 means colomnC
				c2=sheet_obj.cell(row = cell.row, column = 3)
				print(repoList[1])	
				c2.value = repoList[1]
        #column 3 means colomnD
				c3=sheet_obj.cell(row = cell.row, column = 4)
				#print(repoList[1])	
				c3.value = repoList[2]
				wb_obj.save(path)				
			
